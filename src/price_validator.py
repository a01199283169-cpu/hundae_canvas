"""플랫폼 단가표 로드 및 주문 단가 검증."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from src.config_loader import ensure_dirs, load_config, resolve_path
from src.database import connect


@dataclass
class PriceRule:
    platform: str
    frame: str | None
    size: str | None
    width: float | None
    height: float | None
    unit_price: float
    effective_from: str | None = None


def create_price_catalog_template(path: Path | None = None) -> Path:
    """단가표 템플릿 엑셀 생성 (플랫폼 공개 단가 입력용)."""
    config = load_config()
    if path is None:
        path = resolve_path(config["paths"]["price_catalog"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "단가표"

    headers = ["플랫폼", "프레임", "호칭/규격", "가로(mm)", "세로(mm)", "단가", "적용시작일", "비고"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DAEEF3")

    # 샘플 데이터 (실무 입력 참고)
    samples = [
        ["SMT모닝", "박스", "8F", 495, 419, 8300, "2026-01-01", "플랫폼 공개단가"],
        ["프레임연구소", "A", "57", 178, 127, 5390, "2026-01-01", ""],
        ["쇼핑몰 현대", "유화", "8F", 455, 379, 19000, "2026-01-01", ""],
        ["쿠팡", "A", "A3", 420, 297, 16280, "2026-01-01", ""],
    ]
    for row in samples:
        ws.append(row)

    # 컬럼 너비
    widths = [18, 10, 12, 10, 10, 10, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(path)
    wb.close()
    return path


def load_price_catalog(path: Path | None = None) -> list[PriceRule]:
    """단가표 엑셀 읽기."""
    config = load_config()
    if path is None:
        path = resolve_path(config["paths"]["price_catalog"])

    if not path.exists():
        create_price_catalog_template(path)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rules: list[PriceRule] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        platform = str(row[0]).strip()
        frame = str(row[1]).strip() if row[1] else None
        size = str(row[2]).strip() if row[2] else None
        width = float(row[3]) if row[3] not in (None, "") else None
        height = float(row[4]) if row[4] not in (None, "") else None
        unit_price = float(row[5]) if row[5] not in (None, "") else 0.0
        effective = str(row[6]) if row[6] else None

        rules.append(
            PriceRule(
                platform=platform,
                frame=frame,
                size=size,
                width=width,
                height=height,
                unit_price=unit_price,
                effective_from=effective,
            )
        )

    wb.close()
    return rules


def _norm_platform(name: str | None) -> str:
    if not name:
        return ""
    return name.replace("\n", " ").strip()


def _match_rule(
    rules: list[PriceRule],
    platform: str | None,
    frame: str | None,
    size: str | None,
    width: float | None,
    height: float | None,
) -> PriceRule | None:
    """플랫폼+품목 스펙으로 단가 규칙 매칭 (정확 → 부분 일치 순)."""
    plat = _norm_platform(platform)
    best: PriceRule | None = None
    best_score = -1

    for rule in rules:
        score = 0
        if _norm_platform(rule.platform) != plat:
            continue
        score += 10

        if rule.frame and frame and rule.frame == frame:
            score += 5
        elif rule.frame and frame and rule.frame != frame:
            continue

        if rule.size and size and str(rule.size) == str(size):
            score += 3
        if rule.width and width and abs(rule.width - width) < 1:
            score += 2
        if rule.height and height and abs(rule.height - height) < 1:
            score += 2

        if score > best_score:
            best_score = score
            best = rule

    return best if best_score >= 10 else None


def validate_prices(source_file: str | None = None) -> Path:
    """
    DB 주문 품목 vs 단가표 비교 후 검증 리포트 엑셀 생성.
    """
    config = load_config()
    paths = ensure_dirs(config)
    rules = load_price_catalog()
    conn = connect()

    sql = """
        SELECT o.sheet_name, o.order_no, o.platform, o.customer,
               oi.line_no, oi.frame, oi.size, oi.width, oi.height,
               oi.qty, oi.unit_price
        FROM orders o
        JOIN order_items oi ON oi.order_id = o.id
    """
    params: tuple = ()
    if source_file:
        sql += " WHERE o.source_file=?"
        params = (source_file,)
    sql += " ORDER BY o.sheet_date, o.order_no, oi.line_no"

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "단가검증"
    headers = [
        "시트", "주문번호", "플랫폼", "주문자", "품목", "호칭",
        "가로", "세로", "수량", "입력단가", "기준단가", "차액", "상태",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    mismatch_fill = PatternFill("solid", fgColor="FFC7CE")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")

    stats = {"ok": 0, "mismatch": 0, "unregistered": 0}

    for row in rows:
        rule = _match_rule(
            rules,
            row["platform"],
            row["frame"],
            row["size"],
            row["width"],
            row["height"],
        )
        input_price = row["unit_price"] or 0
        ref_price = rule.unit_price if rule else None

        if ref_price is None:
            status = "미등록"
            diff = None
            stats["unregistered"] += 1
            fill = warn_fill
        elif abs(input_price - ref_price) < 1:
            status = "일치"
            diff = 0
            stats["ok"] += 1
            fill = ok_fill
        else:
            status = "불일치"
            diff = input_price - ref_price
            stats["mismatch"] += 1
            fill = mismatch_fill

        ws.append(
            [
                row["sheet_name"],
                row["order_no"],
                row["platform"],
                row["customer"],
                row["frame"],
                row["size"],
                row["width"],
                row["height"],
                row["qty"],
                input_price,
                ref_price,
                diff,
                status,
            ]
        )
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # 요약 시트
    summary = wb.create_sheet("요약")
    summary.append(["항목", "건수"])
    for k, v in stats.items():
        summary.append([k, v])
    summary.append(["검증일시", datetime.now().strftime("%Y-%m-%d %H:%M")])

    out_path = paths["validation"] / f"단가검증_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    wb.close()
    return out_path
