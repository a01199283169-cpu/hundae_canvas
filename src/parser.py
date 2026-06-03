"""모닝프레임 주문내역서 엑셀 파서."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from src.config_loader import load_config
from src.database import connect, init_db, insert_import_log, upsert_order


@dataclass
class ColumnMap:
    """시트별 열 번호 매핑."""

    order_no: int = 2
    date: int = 3
    platform: int = 4
    file: int = 5
    frame: int = 6
    size: int = 7
    width: int = 8
    height: int = 9
    color: int = 10
    plate: int = 11
    acrylic: int = 12
    hook: int = 13
    item_note: int = 14
    qty: int = 15
    customer: int = 16
    phone: int = 17
    address: int = 18
    unit_price: int = 19
    order_qty: int = 20
    sales: int = 21
    deduct: int = 22
    ship: int = 23
    total: int = 24
    pay_card: int = 25
    pay_transfer: int = 26
    pay_bank: int = 27
    remark: int = 28
    expected_ship_type: int = 29
    expected_freight: int = 30
    expected_ship_qty: int = 31
    header_row: int = 6
    data_start_row: int = 9


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    # Postgres DATE 호환 — "2026-06-02 00:00:00" → "2026-06-02"
    m = re.match(r"(\d{4}-\d{2}-\d{2})", text)
    if m:
        return m.group(1)
    return text


def _cell_num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_platform(name: str | None, aliases: dict[str, str]) -> str | None:
    """플랫폼명 줄바꿈·공백 정리 및 별칭 적용."""
    if not name:
        return None
    cleaned = re.sub(r"\s+", " ", name.replace("\n", " ")).strip()
    return aliases.get(name, aliases.get(cleaned, cleaned))


def detect_columns(ws, config: dict) -> ColumnMap:
    """
    헤더 행에서 키워드 기반으로 열 위치 자동 탐지.
    0601처럼 열이 한 칸 밀린 시트도 처리.
    """
    keywords = config["header_keywords"]
    col_map = ColumnMap()
    header_row = 6

    # '구분'과 '단가'가 있는 행을 헤더로 판단
    for r in config["parse"]["header_search_rows"]:
        labels: dict[int, str] = {}
        for c in range(1, 40):
            val = ws.cell(r, c).value
            if val is not None:
                labels[c] = str(val).strip()

        if not any("구분" in v for v in labels.values()):
            continue
        if not any("단가" in v for v in labels.values()):
            continue

        header_row = r
        col_map.header_row = r

        def find_col(keys: list[str], prefer_row: int | None = None) -> int | None:
            search_rows = [prefer_row, r, r + 1, r + 2] if prefer_row else [r, r + 1, r + 2]
            for sr in search_rows:
                if sr is None:
                    continue
                for c, label in labels.items() if sr == r else {
                    c2: str(ws.cell(sr, c2).value).strip()
                    for c2 in range(1, 40)
                    if ws.cell(sr, c2).value is not None
                }.items():
                    for key in keys:
                        if key in label:
                            return c
            return None

        # 6행(또는 탐지행) 라벨 재구성
        labels = {
            c: str(ws.cell(r, c).value).strip()
            for c in range(1, 40)
            if ws.cell(r, c).value is not None
        }
        sub_labels = {
            c: str(ws.cell(r + 2, c).value).strip()
            for c in range(1, 40)
            if ws.cell(r + 2, c).value is not None
        }

        for c, label in labels.items():
            if label in ("NO", "번호"):
                col_map.order_no = c
            elif label == "날짜":
                col_map.date = c
            elif label == "구분":
                col_map.platform = c
            elif label == "파일":
                col_map.file = c
            elif label == "상세내역":
                col_map.frame = c
            elif "주문자" in label:
                col_map.customer = c
            elif label == "연락처":
                col_map.phone = c
            elif label == "주소":
                col_map.address = c
            elif label == "단가":
                col_map.unit_price = c
            elif label == "판매가":
                col_map.sales = c
            elif label == "공제액":
                col_map.deduct = c
            elif label == "택배비":
                col_map.ship = c
            elif "합계" in label:
                col_map.total = c
            elif label == "비고" and c >= col_map.remark:
                col_map.remark = c

        # 7~8행 서브헤더
        for c, label in {
            **{
                c2: str(ws.cell(r + 1, c2).value).strip()
                for c2 in range(1, 40)
                if ws.cell(r + 1, c2).value is not None
            },
            **sub_labels,
        }.items():
            if label == "프레임":
                col_map.frame = c
            elif "호칭" in label:
                col_map.size = c
            elif label == "가로":
                col_map.width = c
            elif label == "세로":
                col_map.height = c
            elif label == "색상":
                col_map.color = c
            elif label == "알판":
                col_map.plate = c
            elif label == "아크릴":
                col_map.acrylic = c
            elif label == "고리":
                col_map.hook = c
            elif label == "비고" and c < col_map.customer:
                col_map.item_note = c
            elif label == "수량" and c < col_map.customer:
                col_map.qty = c
            elif label == "카드":
                col_map.pay_card = c
            elif label in ("계좌이체", "하나", "농협"):
                col_map.pay_transfer = c
            elif label == "무통장":
                col_map.pay_bank = c
            elif label == "타입":
                col_map.expected_ship_type = c
            elif label == "운임비" and c > col_map.ship:
                col_map.expected_freight = c
            elif label == "수량" and c > col_map.remark:
                col_map.expected_ship_qty = c

        # 주문 수량 열: 단가 바로 다음 '수량'
        for c, label in labels.items():
            if label == "수량" and c > col_map.unit_price:
                col_map.order_qty = c
            elif label == "수량" and c < col_map.customer:
                if col_map.qty == 15:  # 기본값이면 덮어쓰기
                    col_map.qty = c

        # NO 헤더(col1)와 날짜(col3) 사이에 빈 열이 있으면 실제 번호는 col2 (0526~0529 형식)
        if col_map.date > col_map.order_no + 1:
            col_map.order_no = col_map.order_no + 1

        col_map.data_start_row = config["parse"]["data_start_row"]
        break

    return col_map


def _read_row_fields(ws, row: int, cols: ColumnMap) -> dict[str, Any]:
    """한 행에서 주문/품목 공통 필드 읽기."""
    return {
        "order_no": _cell_str(ws.cell(row, cols.order_no).value),
        "order_date": _cell_str(ws.cell(row, cols.date).value),
        "platform": _cell_str(ws.cell(row, cols.platform).value),
        "file_ref": _cell_str(ws.cell(row, cols.file).value),
        "frame": _cell_str(ws.cell(row, cols.frame).value),
        "size": _cell_str(ws.cell(row, cols.size).value),
        "width": _cell_num(ws.cell(row, cols.width).value),
        "height": _cell_num(ws.cell(row, cols.height).value),
        "color": _cell_str(ws.cell(row, cols.color).value),
        "plate": _cell_str(ws.cell(row, cols.plate).value),
        "acrylic": _cell_str(ws.cell(row, cols.acrylic).value),
        "hook": _cell_str(ws.cell(row, cols.hook).value),
        "item_note": _cell_str(ws.cell(row, cols.item_note).value),
        "qty": _cell_num(ws.cell(row, cols.qty).value),
        "customer": _cell_str(ws.cell(row, cols.customer).value),
        "phone": _cell_str(ws.cell(row, cols.phone).value),
        "address": _cell_str(ws.cell(row, cols.address).value),
        "unit_price": _cell_num(ws.cell(row, cols.unit_price).value),
        "sales": _cell_num(ws.cell(row, cols.sales).value),
        "deduct": _cell_num(ws.cell(row, cols.deduct).value),
        "ship": _cell_num(ws.cell(row, cols.ship).value),
        "total": _cell_num(ws.cell(row, cols.total).value),
        "pay_card": _cell_str(ws.cell(row, cols.pay_card).value),
        "pay_transfer": _cell_str(ws.cell(row, cols.pay_transfer).value),
        "pay_bank": _cell_str(ws.cell(row, cols.pay_bank).value),
        "remark": _cell_str(ws.cell(row, cols.remark).value),
        "order_qty": _cell_num(ws.cell(row, cols.order_qty).value),
        "expected_ship_type": _cell_str(ws.cell(row, cols.expected_ship_type).value),
        "expected_freight": _cell_str(ws.cell(row, cols.expected_freight).value),
        "expected_ship_qty": _cell_num(ws.cell(row, cols.expected_ship_qty).value),
        "excel_row": row,
    }


def _is_new_order(row_data: dict[str, Any]) -> bool:
    """B열(또는 NO열) 숫자 + 주문 정보 존재 시 새 주문."""
    no = row_data.get("order_no")
    if not no or not str(no).strip().isdigit():
        return False
    # 품목만 있는 행과 구분
    has_order_info = any(
        row_data.get(k)
        for k in ("platform", "customer", "sales", "total", "phone", "file_ref", "address")
    )
    if has_order_info:
        return True
    # NO + 품목만 있어도 신규 주문 (엑셀 첫 행)
    return _has_item_data(row_data)


def _has_item_data(row_data: dict[str, Any]) -> bool:
    """품목 행 여부."""
    return any(
        row_data.get(k)
        for k in ("frame", "unit_price", "qty", "width", "height", "size")
    )


def _extract_sheet_date(ws, sheet_name: str) -> str | None:
    """시트명(MMDD) 또는 작성일 행에서 날짜 추출."""
    for r in range(1, 5):
        val = ws.cell(r, 1).value
        if val and "작성일" in str(val):
            m = re.search(r"(\d{4}-\d{2}-\d{2})", str(val))
            if m:
                return m.group(1)
    # 시트명이 MMDD 형식이면 올해/날짜 조합 (6월 파일 기준)
    if re.match(r"^\d{4}$", sheet_name):
        mm, dd = sheet_name[:2], sheet_name[2:]
        return f"2026-{mm}-{dd}"
    return None


def parse_sheet(ws, sheet_name: str, source_file: str, config: dict) -> list[dict]:
    """시트 1개에서 주문 목록 파싱."""
    cols = detect_columns(ws, config)
    aliases = config.get("platform_aliases", {})
    sheet_date = _extract_sheet_date(ws, sheet_name)
    max_row = config["parse"]["max_data_row"]

    orders: list[dict] = []
    current: dict | None = None

    for row in range(cols.data_start_row, max_row + 1):
        row_data = _read_row_fields(ws, row, cols)

        # 완전 빈 행이면 주문 종료
        if not _has_item_data(row_data) and not _is_new_order(row_data):
            if current:
                orders.append(current)
                current = None
            continue

        if _is_new_order(row_data):
            if current:
                orders.append(current)
            current = {
                "source_file": source_file,
                "sheet_name": sheet_name,
                "sheet_date": sheet_date,
                "order_no": row_data["order_no"],
                "order_date": row_data.get("order_date") or sheet_date,
                "platform": normalize_platform(row_data.get("platform"), aliases),
                "file_ref": row_data.get("file_ref"),
                "customer": row_data.get("customer"),
                "phone": row_data.get("phone"),
                "address": row_data.get("address"),
                "sales": row_data.get("sales"),
                "deduct": row_data.get("deduct"),
                "ship": row_data.get("ship"),
                "total": row_data.get("total"),
                "pay_card": row_data.get("pay_card"),
                "pay_transfer": row_data.get("pay_transfer"),
                "pay_bank": row_data.get("pay_bank"),
                "remark": row_data.get("remark"),
                "order_qty": row_data.get("order_qty"),
                "expected_ship_type": row_data.get("expected_ship_type"),
                "expected_freight": row_data.get("expected_freight"),
                "expected_ship_qty": row_data.get("expected_ship_qty"),
                "start_row": row,
                "items": [],
                "has_image": 0,
            }

        if current and _has_item_data(row_data):
            current["items"].append(
                {
                    "frame": row_data.get("frame"),
                    "size": row_data.get("size"),
                    "width": row_data.get("width"),
                    "height": row_data.get("height"),
                    "color": row_data.get("color"),
                    "plate": row_data.get("plate"),
                    "acrylic": row_data.get("acrylic"),
                    "hook": row_data.get("hook"),
                    "item_note": row_data.get("item_note"),
                    "qty": row_data.get("qty"),
                    "unit_price": row_data.get("unit_price"),
                    "excel_row": row,
                }
            )
            # 후속 행에서 고객·주소·금액 등 보충 (비어 있으면 채움, 있으면 덮어씀)
            for key in (
                "customer", "phone", "address", "platform", "file_ref",
                "sales", "total", "ship", "deduct",
                "order_qty", "expected_ship_type", "expected_freight", "expected_ship_qty",
                "pay_card", "pay_transfer", "pay_bank", "remark",
            ):
                val = row_data.get(key)
                if val is not None and str(val).strip() != "":
                    if key == "platform":
                        current[key] = normalize_platform(val, aliases)
                    else:
                        current[key] = val

    if current:
        orders.append(current)

    return orders


def parse_workbook(xlsx_path: Path) -> dict[str, list[dict]]:
    """엑셀 파일 전체 시트 파싱."""
    config = load_config()
    source_file = xlsx_path.name
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result[sheet_name] = parse_sheet(ws, sheet_name, source_file, config)

    wb.close()
    return result


def import_to_db(xlsx_path: Path) -> dict[str, Any]:
    """엑셀 파싱 후 SQLite 적재."""
    init_db()
    conn = connect()
    parsed = parse_workbook(xlsx_path)
    summary = {"sheets": {}, "total_orders": 0, "total_items": 0}

    for sheet_name, orders in parsed.items():
        import_id = insert_import_log(
            conn,
            xlsx_path.name,
            sheet_name,
            len(orders),
            sum(len(o["items"]) for o in orders),
        )
        for order in orders:
            upsert_order(conn, order, import_id)

        item_count = sum(len(o["items"]) for o in orders)
        summary["sheets"][sheet_name] = {"orders": len(orders), "items": item_count}
        summary["total_orders"] += len(orders)
        summary["total_items"] += item_count

    conn.close()
    summary["source_file"] = xlsx_path.name
    return summary
