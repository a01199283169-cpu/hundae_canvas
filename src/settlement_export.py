"""일자별 매출현황 Excel 생성."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from src.config_loader import ensure_dirs, load_config
from src.web_service import get_settlement_data


def _num(val) -> float:
    try:
        return float(val or 0)
    except (TypeError, ValueError):
        return 0.0


def export_monthly_settlement(
    month: str | None = None,
    *,
    period: str = "month",
    day: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    platform: str | None = None,
    pay_method: str | None = None,
    payment_status: str | None = None,
    source_file: str | None = None,
) -> Path:
    """조회 조건에 맞는 매출현황 Excel 생성."""
    config = load_config()
    paths = ensure_dirs(config)

    data = get_settlement_data(
        period=period,
        month=month,
        day=day,
        date_from=date_from,
        date_to=date_to,
        platform=platform,
        pay_method=pay_method,
        payment_status=payment_status,
    )
    orders = data["orders"]
    daily_stats = data["daily_stats"]
    daily_rows = data["daily_rows"]
    grand = data["grand"]
    sales_rows = data["sales_rows"]

    if source_file:
        orders = [o for o in orders if o.get("source_file") == source_file]

    wb = openpyxl.Workbook()

    # --- 일자별 요약 시트 ---
    ws_sum = wb.active
    ws_sum.title = "일자별합계"
    label_parts = []
    if period == "day" and day:
        label_parts.append(f"일자 {day}")
    elif period == "range" and date_from and date_to:
        label_parts.append(f"기간 {date_from}~{date_to}")
    elif month:
        label_parts.append(month)
    ym_label = " · ".join(label_parts) if label_parts else "전체"
    ws_sum.append([f"모닝프레임 매출현황 ({ym_label})"])
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.append([])

    sum_headers = [
        "일자", "주문건수", "판매가합계", "공제액합계", "택배비합계",
        "합계금액", "결제완료", "미결",
    ]
    ws_sum.append(sum_headers)
    for cell in ws_sum[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DAEEF3")

    for row in daily_rows:
        if row["type"] == "month_header":
            ws_sum.append([row["label"]])
            for cell in ws_sum[ws_sum.max_row]:
                cell.font = Font(bold=True, color="6366F1")
        elif row["type"] == "month_subtotal":
            ws_sum.append([
                f"▸ {row['label']} 합계", row["order_count"], row["sales"], row["deduct"],
                row["ship"], row["total"], row["completed_count"], row["pending_count"],
            ])
            for cell in ws_sum[ws_sum.max_row]:
                cell.font = Font(bold=True)
        else:
            ws_sum.append([
                row["date"], row["order_count"], row["sales"], row["deduct"], row["ship"],
                row["total"], row["completed_count"], row["pending_count"],
            ])

    ws_sum.append([])
    ws_sum.append([
        "합계", grand["order_count"], grand["sales"], grand["deduct"],
        grand["ship"], grand["total"], grand["completed_count"], grand["pending_count"],
    ])
    for cell in ws_sum[ws_sum.max_row]:
        cell.font = Font(bold=True)

    # --- 매출 상세 시트 (일 소계·월 구분 포함) ---
    ws_all = wb.create_sheet("매출상세")
    detail_headers = [
        "NO", "일자", "구분", "상세내역", "주문자", "단가", "수량",
        "판매가", "공제액", "택배비", "합계", "결제수단", "결제여부", "입금일", "비고",
    ]
    ws_all.append(detail_headers)
    for cell in ws_all[1]:
        cell.font = Font(bold=True)

    for row in sales_rows:
        if row["type"] == "month_header":
            ws_all.append([row["label"]])
            ws_all[ws_all.max_row][0].font = Font(bold=True, color="6366F1")
        elif row["type"] == "day_divider":
            ws_all.append([f"── {row['label']} ──"])
            ws_all[ws_all.max_row][0].font = Font(bold=True, color="64748B")
        elif row["type"] == "month_subtotal":
            ws_all.append([
                "", f"▸ {row['label']} 합계 ({row['order_count']}건)", "", "", "",
                "", "", row["sales"], row["deduct"], row["ship"], row["total"],
                "", "", "", "",
            ])
            for cell in ws_all[ws_all.max_row]:
                cell.font = Font(bold=True)
        elif row["type"] == "day_subtotal":
            ws_all.append([
                "", row["date"], f"일 합계 ({row['order_count']}건)", "", "",
                "", "", row["sales"], row["deduct"], row["ship"], row["total"],
                "", "", "", "",
            ])
            for cell in ws_all[ws_all.max_row]:
                cell.font = Font(bold=True)
        else:
            o = row["order"]
            ws_all.append([
                o.get("order_no"),
                o.get("display_date"),
                o.get("platform"),
                o.get("spec_summary"),
                o.get("customer"),
                o.get("unit_price"),
                o.get("order_qty") or o.get("item_qty"),
                o.get("sales"),
                o.get("deduct"),
                o.get("ship"),
                o.get("total"),
                o.get("payment_label"),
                o.get("payment_status_label"),
                o.get("deposit_date"),
                o.get("remark"),
            ])

    ws_verify = wb.create_sheet("검증")
    ws_verify.append(["검증 항목", "값"])
    ws_verify.append(["주문 건수", len(orders)])
    ws_verify.append(["일자 수", len(daily_stats)])
    ws_verify.append(["판매가 합계", grand["sales"]])
    ws_verify.append(["합계금액 합계", grand["total"]])
    ws_verify.append(["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    ym_str = (month or ym_label).replace("-", "")[:8]
    out_path = paths["settlement"] / f"매출현황_{ym_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    wb.close()
    return out_path
