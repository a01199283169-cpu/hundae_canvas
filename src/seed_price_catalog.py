"""주문 데이터에서 단가표 초기 샘플 생성 (플랫폼 공개단가 입력 보조)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from src.config_loader import load_config, resolve_path
from src.database import connect


def seed_price_catalog_from_orders() -> Path:
    """DB 주문 품목에서 고유 (플랫폼+프레임+규격) 조합 추출해 단가표에 추가."""
    config = load_config()
    path = resolve_path(config["paths"]["price_catalog"])
    path.parent.mkdir(parents=True, exist_ok=True)

    conn = connect()
    rows = conn.execute(
        """
        SELECT DISTINCT o.platform, oi.frame, oi.size, oi.width, oi.height, oi.unit_price
        FROM orders o
        JOIN order_items oi ON oi.order_id = o.id
        WHERE o.platform IS NOT NULL AND oi.unit_price IS NOT NULL
        ORDER BY o.platform, oi.frame, oi.size
        """
    ).fetchall()
    conn.close()

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "단가표"
        headers = ["플랫폼", "프레임", "호칭/규격", "가로(mm)", "세로(mm)", "단가", "적용시작일", "비고"]
        ws.append(headers)

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            existing.add(tuple(str(x) if x is not None else "" for x in row[:6]))

    added = 0
    for row in rows:
        key = (
            str(row["platform"] or ""),
            str(row["frame"] or ""),
            str(row["size"] or ""),
            str(row["width"] or ""),
            str(row["height"] or ""),
            str(row["unit_price"] or ""),
        )
        if key in existing:
            continue
        ws.append(
            [
                row["platform"],
                row["frame"],
                row["size"],
                row["width"],
                row["height"],
                row["unit_price"],
                "2026-01-01",
                "주문 데이터에서 자동 추출 (플랫폼 공개단가로 확인 후 수정)",
            ]
        )
        added += 1

    wb.save(path)
    wb.close()
    return path


if __name__ == "__main__":
    p = seed_price_catalog_from_orders()
    print(f"단가표 업데이트: {p}")
