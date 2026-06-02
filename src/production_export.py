"""생산지시서 엑셀 자동 생성."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.config_loader import ensure_dirs, load_config
from src.database import connect, fetch_order_items


def _thin_border() -> Border:
    side = Side(style="thin", color="999999")
    return Border(left=side, right=side, top=side, bottom=side)


def export_production_sheet(
    source_file: str | None = None,
    sheet_filter: str | None = None,
) -> Path:
    """
    생산지시서 엑셀 생성.
    현장 프린트용 간결 형식.
    """
    config = load_config()
    paths = ensure_dirs(config)
    conn = connect()

    sql = """
        SELECT o.*,
               (SELECT GROUP_CONCAT(oi.image_file, '; ')
                FROM order_images oi WHERE oi.order_id=o.id AND oi.mapped=1) AS image_paths
        FROM orders o
        WHERE 1=1
    """
    params: list = []
    if source_file:
        sql += " AND o.source_file=?"
        params.append(source_file)
    if sheet_filter:
        sql += " AND o.sheet_name=?"
        params.append(sheet_filter)
    sql += " ORDER BY o.sheet_date, o.sheet_name, CAST(o.order_no AS INTEGER)"

    orders = conn.execute(sql, params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "생산지시서"

    title = f"모닝프레임 생산지시서 ({datetime.now().strftime('%Y-%m-%d')})"
    ws.merge_cells("A1:P1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "작업순번", "시트", "주문번호", "플랫폼", "프레임", "호칭",
        "가로(mm)", "세로(mm)", "색상", "알판", "아크릴", "고리",
        "수량", "이미지", "주문자", "연락처", "배송지", "비고",
    ]
    ws.append([])  # 빈 행
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    seq = 0
    for order in orders:
        items = fetch_order_items(conn, order["id"])
        for item in items:
            seq += 1
            has_img = "Y" if order["has_image"] else ""
            img_path = order["image_paths"] or ""

            spec = f"{item['width'] or ''}x{item['height'] or ''}"
            if item["size"]:
                spec = str(item["size"])

            ws.append(
                [
                    seq,
                    order["sheet_name"],
                    order["order_no"],
                    order["platform"],
                    item["frame"],
                    item["size"],
                    item["width"],
                    item["height"],
                    item["color"],
                    item["plate"],
                    item["acrylic"],
                    item["hook"],
                    item["qty"],
                    has_img,
                    order["customer"],
                    order["phone"],
                    (order["address"] or "")[:80],
                    order["remark"] or item["item_note"] or "",
                ]
            )
            row_idx = ws.max_row
            if has_img and img_path:
                ws.cell(row_idx, 14).value = f"Y ({Path(img_path).name})"

            for cell in ws[row_idx]:
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 열 너비
    widths = [8, 8, 8, 14, 10, 8, 8, 8, 10, 8, 8, 8, 6, 10, 12, 14, 40, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    # 일별 시트도 생성
    sheet_names = sorted({o["sheet_name"] for o in orders})
    for sn in sheet_names:
        daily = wb.create_sheet(f"생산_{sn}")
        daily.append([f"생산지시 - {sn}"])
        daily.append(headers)
        seq_d = 0
        for order in orders:
            if order["sheet_name"] != sn:
                continue
            items = fetch_order_items(conn, order["id"])
            for item in items:
                seq_d += 1
                daily.append(
                    [
                        seq_d,
                        order["sheet_name"],
                        order["order_no"],
                        order["platform"],
                        item["frame"],
                        item["size"],
                        item["width"],
                        item["height"],
                        item["color"],
                        item["plate"],
                        item["acrylic"],
                        item["hook"],
                        item["qty"],
                        "Y" if order["has_image"] else "",
                        order["customer"],
                        order["phone"],
                        (order["address"] or "")[:80],
                        order["remark"] or item["item_note"] or "",
                    ]
                )

    conn.close()

    suffix = sheet_filter or "전체"
    out_path = paths["production"] / f"생산지시서_{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    wb.close()
    return out_path
